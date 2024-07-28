import requests
import json
import openpyxl


def test_add_multiple_students():
    #API
    api_url = "https://thetestingworldapi.com/api/studentsDetails"
    file = open('C:\\Users\\Marina_my_folder\\testing_purpose\\AddNewStudent.json', 'r')
    json_request = json.loads(file.read())
    #Excel code
    vk = openpyxl.load_workbook('C:\\Users\\Marina_my_folder\\testing_purpose\\TestData.xlsx')
    sh = vk['Лист1']
    rows = sh.max_row

    for i in range(2,rows+1):
        cell_first_name = sh.cell(row=i, column=1)
        cell_mid_name = sh.cell(row=i, column=2)
        cell_last_name = sh.cell(row=i, column=3)
        cell_db = sh.cell(row=i, column=4)
        json_request['first_name']= cell_first_name.value
        json_request['middle_name']= cell_mid_name.value
        json_request['last_name']= cell_last_name.value
        json_request['date_of_birth']= cell_db.value

        response = requests.post(api_url, json_request)
        print(response.text)
        print(response.status_code)
        assert response.status_code == 201
